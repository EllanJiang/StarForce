import os
import openpyxl
from .util import *
import time
showkey_title = "showkey"
state_title = "state"
comment_title = "comment"
content_title = "content"

exception_info = []
def start_merge():
    print("start merge localization excel")
    current_path = os.getcwd()
    source_folder_path = os.path.join(current_path, "../LocalizationKeyExcel")
    excel_files = os.listdir(source_folder_path)
    keys = {}
    origin_keys = {}
    origin_path = ""
    excels_path = []
    for file_name in excel_files:
        path = os.path.join(source_folder_path, file_name)
        if os.path.isfile(path):
            if contain_chinese(file_name) or not (".xlsx" == os.path.splitext(file_name)[-1]):
                print("ignore excel" + file_name)
                continue
            # print(file_name)
            if file_name == "Config.xlsx":
                origin_keys = read_excel(path)
                origin_path = path
            else:
                excels_path.append(path)
                keys.update({file_name: read_excel(path)})

    _check_duplication(keys)
    if len(exception_info)>0:
        for i in range(len(exception_info)):
            print(exception_info[i])
        raise Exception("-----------打表错误----------------")
    workbook = openpyxl.load_workbook(origin_path)
    sheet_names = workbook.sheetnames
    work_sheet = workbook[sheet_names[0]]
    rows = work_sheet.max_row
    current_row = rows
    for k, v in keys.items():
        for key, value in v.items():
            if key not in origin_keys:
                current_row = current_row + 1
                old_staff_record = work_sheet.cell(current_row, 4).value
                work_sheet.cell(current_row, 1).value = key
                work_sheet.cell(current_row, 2).value = value[0]
                if value[1] is not None:
                    work_sheet.cell(current_row, 3).value = value[1]
                # if old_staff_record is None or old_staff_record == "":
                #     work_sheet.cell(current_row, 4).value = k
                # else:
                #     work_sheet.cell(current_row, 4).value = str(old_staff_record) + "," + k
                work_sheet.cell(current_row, 4).value = k
                work_sheet.cell(current_row,5).value = time.strftime("%Y-%m-%d", time.localtime())

                origin_keys.update({key: value})
            else:
                row = origin_keys[key][2]
                old_value = work_sheet.cell(row, 2).value
                old_staff_record = work_sheet.cell(row, 4).value
                work_sheet.cell(row, 1).value = key
                work_sheet.cell(row, 2).value = value[0]
                work_sheet.cell(row, 4).value = k
                work_sheet.cell(row, 5).value = time.strftime("%Y-%m-%d", time.localtime())
                work_sheet.cell(row, 7).value = old_value
                # if value[1] is not None:
                #     work_sheet.cell(row, 3).value = value[1]
                # if old_staff_record is None or old_staff_record == "":
                #     work_sheet.cell(row, 4).value = k
                # else:
                #     work_sheet.cell(row, 4).value = str(old_staff_record) + "," + k


    for i, val in enumerate(excels_path):
        excel_workbook = openpyxl.load_workbook(val)
        sheet_names = excel_workbook.sheetnames
        work_sheet = excel_workbook[sheet_names[0]]
        rows = work_sheet.max_row
        title_dic = {}
        for col in range(1, 6):
            v = work_sheet.cell(1, col).value
            if v != "" and v is not None:
                title_dic.update({str(v): int(col)})
        not_delete_count = 0
        for i in range(2, rows + 1):
            if state_title in title_dic:
                state = str(work_sheet.cell(2+not_delete_count, title_dic[state_title]).value)
                if state == "准备中":
                    not_delete_count = not_delete_count +1
                    continue
            work_sheet.delete_rows(2+not_delete_count)
        excel_workbook.save(val)
        excel_workbook.close()

    workbook.save(origin_path)
    workbook.close()
    print("end merge localization excel")


def _check_duplication(keys):
    dic = {}
    for k, v in keys.items():
        for key, value in v.items():
            if(key in dic):
                format = "{0}和{1}存在重复的key={2}"
                exception_info.append(format.format(k,dic[key],key))
                # raise Exception(format.format(k,dic[key],key))
            dic.update({key:k})

def get_incremental(folder_path, excels_path):
    excel_files = os.listdir(folder_path)
    for file_name in excel_files:
        path = os.path.join(folder_path, file_name)
        if os.path.isfile(path):
            excels_path.append(path)
            return read_excel(path)


def read_excel(file_path):
    print("read excel "+file_path)
    workbook = openpyxl.load_workbook(file_path)
    sheet_names = workbook.sheetnames
    work_sheet = workbook[sheet_names[0]]
    rows = work_sheet.max_row
    column = work_sheet.max_column
    dic = {}
    title_dic = {}
    for col in range(1, 6):
        v = work_sheet.cell(1, col).value
        if v != "" and v is not None:
            title_dic.update({str(v): int(col)})

    for row in range(2, rows + 1):
        if work_sheet.cell(row, 1).value != "" and work_sheet.cell(row, 1).value is not None:
            if state_title in title_dic:
                state = str(work_sheet.cell(row, title_dic[state_title]).value)
                if state == "准备中":
                    continue
            show_key = work_sheet.cell(row, title_dic[showkey_title]).value
            content = work_sheet.cell(row, title_dic[content_title]).value
            if show_key is None or show_key == "":
                continue
            if content is None or content == "":
                exception_info.append("表格存在内容为空字段 excel="+file_path)
                # raise Exception("表格存在内容为空字段 excel="+file_path)
            comment = None
            if comment_title in title_dic:
                comment = work_sheet.cell(row, title_dic[comment_title]).value
            show_key = str(show_key).strip()
            show_key = show_key.replace('\n' , '').replace('\r','')
            if show_key in dic:
                error_info = "key重复 key=" + show_key + " excel=" + file_path
                exception_info.append(error_info)
                # raise Exception(error_info)
            else:
                dic.update({show_key: [content, comment,row]})
    workbook.close()
    return dic


if __name__ == "__main__":
    start_merge()

