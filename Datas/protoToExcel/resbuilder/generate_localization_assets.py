import sys
import os
import openpyxl
from unityparser import UnityDocument
import shutil
import base64
from .util import *


showkey_title = "showkey"
state_title = "state"
comment_title = "comment"
content_title = "content"

def _write_asset(language, sheet_name, save_path, template_path):
    try:
        # 生成Asset格式语言包
        if len(language) > 0:
            if sheet_name == "config" or sheet_name == "assetupdate":
                language_asset_file_name = "tbl_" + sheet_name.lower() + "_language.asset"
            else:
                language_asset_file_name = sheet_name.lower() + ".asset"
            if not os.path.exists(save_path):
                return
            language_asset_name = os.path.join(save_path, language_asset_file_name)
            src_asset_path = template_path
            print(src_asset_path)
            shutil.copyfile(src_asset_path, language_asset_name)
            doc = UnityDocument.load_yaml(language_asset_name)
            ProjectSettings = doc.entry
            if sheet_name == "config" or sheet_name == "assetupdate":
                ProjectSettings.m_Name = "tbl_" + sheet_name.lower() + "_language"
            else:
                ProjectSettings.m_Name = sheet_name.lower()

            arrays = []
            for (k, v) in language.items():
                dic = {}
                dic['showkey'] = k
                dic['key'] = get_str_md5_9(k)
                dic['type'] = 0
                value_dic = {}
                sb = bytes(v, encoding='utf-8')
                if ("<测试>" in v) or ("<test>" in v):
                    continue
                str_result = base64.b64encode(sb)
                value_dic['dataValue'] = str(str_result, encoding="utf-8")
                dic['value'] = value_dic
                arrays.append(dic)
            ProjectSettings.Source['entries'] = arrays
            doc.dump_yaml(language_asset_name)
    except BaseException as e:
        print("write excel item by asset error:(%s)" % sheet_name)
        raise


def _ReadKeys(source_folder_path):
    dic = {}
    excel_files = os.listdir(source_folder_path)
    for file_name in excel_files:
        path = os.path.join(source_folder_path, file_name)
        if os.path.isfile(path):
            if contain_chinese(file_name) or not (".xlsx" == os.path.splitext(file_name)[-1]):
                print("ignore excel" + file_name)
                continue
            if file_name == "Config.xlsx":
                print("ignore excel" + file_name)
                continue
            workbook = openpyxl.load_workbook(path)
            sheet_names = workbook.sheetnames
            work_sheet = workbook[sheet_names[0]]
            rows = work_sheet.max_row

            title_dic = {}
            for col in range(1, 6):
                v = work_sheet.cell(1, col).value
                if v != "" and v is not None:
                    title_dic.update({str(v): int(col)})

            for row in range(2, rows + 1):
                # if state_title in title_dic:
                #     state = str(work_sheet.cell(row, title_dic[state_title]).value)
                    # if state == "准备中":
                    #     continue
                show_key = work_sheet.cell(row, title_dic[showkey_title]).value
                content = work_sheet.cell(row, title_dic[content_title]).value
                if show_key is not None and show_key != "":
                    if show_key in dic:
                        error_format = "{0}和{1}存在重复的key={2}"
                        raise Exception(error_format.format(file_name, dic[show_key][1], str(show_key)))
                        print("key重复" + str(show_key))
                    else:
                        dic.update({str(show_key): [str(content), file_name]})

    return dic

def _ReadConfigKeys(source_folder_path):
    dic = {}
    path = os.path.join(source_folder_path,"Config.xlsx")
    if os.path.isfile(path):
        workbook = openpyxl.load_workbook(path)
        sheet_names = workbook.sheetnames
        work_sheet = workbook[sheet_names[0]]
        rows = work_sheet.max_row
        for row in range(2, rows + 1):
            show_key = work_sheet.cell(row, 1).value
            content = work_sheet.cell(row, 2).value
            if show_key is not None and show_key != "":
                if show_key in dic:
                    error_format = "{0}和{1}存在重复的key={2}"
                    raise Exception(error_format.format(file_name, dic[show_key][1], str(show_key)))
                    print("key重复" + str(show_key))
                else:
                    dic.update({str(show_key): [str(content), "Config.xlsx"]})
    return dic


class GenerateLocalizationAssets:

    def __init__(self, inputs,save_path, template_path, source_folder_path):
        if save_path == "" or save_path is None:
            self.save_path = os.path.join(os.getcwd(), "../", "out/excel/language_asset_out")
        else:
            self.save_path = save_path
        if template_path == "":
            self.template_path = os.path.join(os.getcwd(), "../", 'template.asset')
        else:
            self.template_path = template_path
        if source_folder_path == "":
            source_folder_path = os.path.join(os.getcwd(), "../LocalizationKeyExcel")

        lookup_json_file = load_config_file("excel/BattleConfigLookup.json")
        toclient_sheets = lookup_json_file["toClient"]

        all_keys = {}
        all_keys_exclude_config = _ReadKeys(source_folder_path)
        config_keys = _ReadConfigKeys(source_folder_path)
        for k,v in config_keys.items():
            all_keys.update({k:v})

        for k, v in all_keys_exclude_config.items():
            all_keys.update({k: v})

        # for k, v in all_keys.items():
        #     print(k,v)
        all_used_keys = {}
        self.config_keys = {}
        for filepath in inputs:
            xls_file_name = os.path.splitext(os.path.basename(filepath))[0]
            to_client = (xls_file_name.lower() in toclient_sheets)
            if not to_client:
                print("Generate Asset SKIP: file (%s) name (%s)" % (filepath,xls_file_name))
                continue
            workbook = openpyxl.load_workbook(filepath , data_only=True)
            sheet_names = workbook.sheetnames
            for name in sheet_names:
                if contain_chinese(name):
                    print("SKIP: file (%s) sheet (%s)" % (filepath, name))
                    continue
                work_sheet = workbook[name]
                rows = work_sheet.max_row
                columns = work_sheet.max_column
                for column in range(1,columns+1):
                    if work_sheet.cell(1, column).value != "string":
                        continue
                    for row in range(4, rows+1):
                        show_key = str(work_sheet.cell(row, column).value)
                        str_md5 = get_str_md5_9(show_key)
                        if show_key.startswith("C_") or show_key.startswith("dp_C_"):
                            all_used_keys.update({show_key: str_md5})

        self.hotfix_keys = {}
        for key,value in all_keys.items():
            if key.startswith("C_AssetUpdate_msg"):
                self.hotfix_keys.update({key: value[0]})
            else:
                self.config_keys.update({key: value[0]})
            #筛选使用的key
            # if key in all_used_keys:
            #     self.config_keys.update({key:value[0]})


    def GenerateAssets(self):
        print("start generate localization asset")
        _write_asset(self.config_keys, "config", self.save_path, self.template_path)
        _write_asset(self.hotfix_keys, "assetupdate", self.save_path, self.template_path)
        # _write_asset(self.prefab_keys, "prefab", self.save_path, self.template_path)
        print("end generate localization asset")


# if __name__ == "__main__":
#     f = GenerateLocalizationAssets("", "", "")
#     f.GenerateAssets()

