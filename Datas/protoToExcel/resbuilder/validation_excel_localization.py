import os
import xlwings
import openpyxl
# from .util import *
from util import *
import threading
import time

def validate():
    current_path = os.getcwd()
    source_folder_path = os.path.join(current_path, "../excel")
    excel_files = os.listdir(source_folder_path)
    lookup_json_file = load_config_file("../excel/BattleConfigLookup.json")
    toclient_sheets = lookup_json_file["toClient"]
    invalid_keys_dic = {}
    for file_name in excel_files:
        file_path = os.path.join(source_folder_path, file_name)
        xls_file_name = os.path.splitext(os.path.basename(file_path))[0]
        to_client = (xls_file_name.lower() in toclient_sheets)
        if not to_client:
            print("Generate Asset SKIP: file (%s) name (%s)" % (file_path, xls_file_name))
            continue
        if contain_chinese(file_name) or not (".xlsx" == os.path.splitext(file_name)[-1]):
            print("ignore excel" + file_name)
        else:
            is_excel_data_dirty = False
            workbook = openpyxl.load_workbook(file_path)
            sheet_names = workbook.sheetnames
            work_sheet = workbook[sheet_names[0]]
            rows = work_sheet.max_row
            column = work_sheet.max_column
            invalid_keys = []
            for col in range(1, work_sheet.max_column + 1):
                if work_sheet.cell(1, col).value == "string" and work_sheet.cell(2, col).value != "id":
                    for row in range(4, work_sheet.max_row + 1):
                        key = work_sheet.cell(row, col).value
                        if key is None:
                            continue
                        key = str(key)
                        if not key.startswith("C_") and not key.startswith("dp_C_"):
                            is_excel_data_dirty = True
                            invalid_keys.append(key)

            if is_excel_data_dirty:
                if file_name != "DirtyWords.xlsx":
                    invalid_keys_dic.update({file_name: invalid_keys})

    for k,v in invalid_keys_dic.items():
        print(k,v)


if __name__ == "__main__":
    validate()