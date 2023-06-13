import os
import openpyxl
from unityparser import UnityDocument
import shutil
import base64
from util import *

def generate_keys():
    lookup_json_file = load_config_file("../excel/BattleConfigLookup.json")
    toclient_sheets = lookup_json_file["toClient"]
    current_path = os.getcwd()
    source_folder_path = os.path.join(current_path, "../excel")
    excel_files = os.listdir(source_folder_path)
    all_keys = {}
    for file_name in excel_files:
        file_path = os.path.join(source_folder_path, file_name)
        xls_file_name = os.path.splitext(os.path.basename(file_path))[0]
        to_client = (xls_file_name.lower() in toclient_sheets)
        if not to_client:
            print("not client excel" + file_name)
            continue
        if contain_chinese(file_name) or not (".xlsx" == os.path.splitext(file_name)[-1]):
            print("ignore excel" + file_name)
        else:
            # print(file_path)
            workbook = openpyxl.load_workbook(file_path)
            sheet_names = workbook.sheetnames
            work_sheet = workbook[sheet_names[0]]
            rows = work_sheet.max_row
            column = work_sheet.max_column
            for col in range(1, work_sheet.max_column + 1):
                if work_sheet.cell(1, col).value == "string":
                    for row in range(4, work_sheet.max_row + 1):
                        key = str(work_sheet.cell(row, col).value)
                        if key == "" or key is None:
                            continue
                        if ("<测试>" in key) or ("<test>" in key):
                            continue
                        if "dp_C_" in key:
                            continue
                        show_key = "dp_C_" + key
                        work_sheet.cell(row, col).value = show_key
                        all_keys.update({show_key: key})
            workbook.save(file_path)

    print(len(all_keys))
    for (k, v) in all_keys.items():
        print(k)




if __name__ == "__main__":
    generate_keys()