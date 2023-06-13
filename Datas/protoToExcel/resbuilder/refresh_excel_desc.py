import os
import xlwings
import openpyxl
# from .util import *
from util import *
import threading
import time

desc_name = "多语言注释"

# from win32com.client import Dispatch
# def just_open(filename):
#     print(filename)
#     xlApp = Dispatch("Excel.Application")
#     xlApp.Visible = False
#     xlBook = xlApp.Workbooks.Open(filename)
#     xlBook.Save()
#     xlBook.Close()


def resave_all_excels(dirty_excels):
    current_path = os.getcwd()
    source_folder_path = os.path.join(current_path, "../excel")
    excel_files = os.listdir(source_folder_path)
    excel_app = xlwings.App(visible=False)
    print(len(dirty_excels))
    for file_name in excel_files:
        file_path = os.path.join(source_folder_path, file_name)
        if contain_chinese(file_name) or not (".xlsx" == os.path.splitext(file_name)[-1]):
            print("ignore excel" + file_name)
        else:
            print(file_path)
            if file_name in dirty_excels:
                excel_book = excel_app.books.open(file_path)
                excel_book.save()
                excel_book.close()

    excel_app.quit()


def read_keys():
    current_path = os.getcwd()
    source_folder_path = os.path.join(current_path, "../LocalizationKeyExcel")
    file_path = os.path.join(source_folder_path, "Config.xlsx")
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet_names = workbook.sheetnames
    work_sheet = workbook[sheet_names[0]]
    rows = work_sheet.max_row
    column = work_sheet.max_column
    dic = {}
    for row in range(2, rows + 1):
        show_key = str(work_sheet.cell(row, 1).value)
        content = str(work_sheet.cell(row, 2).value)
        if show_key in dic:
            print("key重复" + show_key)
        else:
            dic.update({show_key: content})
    return dic


def refresh_excels():
    current_path = os.getcwd()
    source_folder_path = os.path.join(current_path, "../excel")
    excel_files = os.listdir(source_folder_path)
    formula_excel = ["AchievementGoal.xlsx","BattleBuyLimitConfig.xlsx","BattleInitWeaponConfig.xlsx","BattleMoneyMatchEventConfig.xlsx","BattleMoneyReward.xlsx","BattlePropsConfig.xlsx","CaseConfig.xlsx","EntityProxyModeConfig.xlsx","EquipBagConfig.xlsx","EquipBattleConfig.xlsx","GameModeBaseConfig.xlsx","GameModeDamageRateConfig.xlsx","GameModeHudConfig.xlsx","GameModeLifeConfig.xlsx","GMInstruction.xlsx","GroupItemsShop.xlsx","InnerGuideConfig.xlsx","InnerGuideTipsConfig.xlsx","ItemConfig.xlsx","ItemInitConfig.xlsx","ItemResource.xlsx","ItemRewardsConfig.xlsx","ItemUseConfig.xlsx","KillNotifyConfig.xlsx","KillNotifyItemConfig.xlsx","LuckyStrikeCase.xlsx","LuckyStrikeConfig.xlsx","LuckyStrikeCosts.xlsx","LuckyStrikePurchase.xlsx","LuckyStrikeReward.xlsx","LuckyStrikeRule.xlsx","MatchRankLevel.xlsx","MaterialDataConfig.xlsx","MonsterLevelData.xlsx","MusicBoxConfig.xlsx","MusicBoxItemConfig.xlsx","MusicBoxKeyConfig.xlsx","MusicBoxModelConfig.xlsx","PaintConfig.xlsx","PaintTypeConfig.xlsx","PlayerHangUpConfig.xlsx","RecommendBuyConfig.xlsx","RoleVoxData.xlsx","SceneModeConfig.xlsx","ShopShowItems.xlsx","SigninConfig.xlsx","SigninRepairCosts.xlsx","SkillDataConfig.xlsx","TacticalVoice.xlsx","TaskCondition.xlsx","TaskConfig.xlsx","WeaponDataConfig.xlsx","WeaponDetailConfig.xlsx","WeaponGroupCard.xlsx","WeaponPaintConfig.xlsx","WeaponPaintKit.xlsx"]
    dirty_excels = []
    for file_name in excel_files:
        file_path = os.path.join(source_folder_path, file_name)
        if contain_chinese(file_name) or not (".xlsx" == os.path.splitext(file_name)[-1]):
            print("ignore excel" + file_name)
        else:
            is_excel_data_dirty = False
            workbook = openpyxl.load_workbook(file_path)
            sheet_names = workbook.sheetnames
            print(sheet_names[0])
            work_sheet = workbook[sheet_names[0]]
            rows = work_sheet.max_row
            column = work_sheet.max_column
            # strs = []
            # for col in range(1, column + 1):
            #     if work_sheet.cell(1, col).value == "string" and work_sheet.cell(2, col).value != "id":
            #         if work_sheet.cell(3, col + 1).value == desc_name:
            #             print("-----")
            #         else:
            #             if file_name in formula_excel:
            #                 print("存在公式excel"+file_name)
            #             else:
            #                 strs.append(col)
            #
            # for index in range(len(strs)):
            #     work_sheet.insert_cols(strs[len(strs) - index - 1] + 1)
            #     work_sheet.cell(3, strs[len(strs) - index - 1] + 1).value = desc_name
            keys_dic = read_keys()
            for col in range(1, work_sheet.max_column + 1):
                # if(work_sheet.cell(4, col).data_type == "f"):
                #     formula_str = formula_str+ "\"" +file_name +"\"" +",";
                #     break
                if work_sheet.cell(1, col).value == "string" and work_sheet.cell(2, col).value != "id":
                    if work_sheet.cell(3, col + 1).value == desc_name:
                        for row in range(4, work_sheet.max_row + 1):
                            key = str(work_sheet.cell(row, col).value)
                            if key in keys_dic and work_sheet.cell(row, col + 1).value!=keys_dic[key]:
                                is_excel_data_dirty = True
                                work_sheet.cell(row, col + 1).value = keys_dic[key]

            if is_excel_data_dirty:
                print("excel changed name="+file_name)
                dirty_excels.append(file_name)
                workbook.save(file_path)

    time.sleep(2)
    resave_all_excels(dirty_excels)


if __name__ == "__main__":
    refresh_excels()


