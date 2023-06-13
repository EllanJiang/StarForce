#!/usr/bin/python
# -*- coding:utf-8 _*-

import os
import sys
import openpyxl
from util import *
import xlwings

exception_info = []

def start_replace():
    print("start_replace")
    print(sys.getdefaultencoding())
    current_path = os.getcwd()
    source_folder_path = os.path.join(current_path, "../LocalizationKeyExcel")
    excel_files = os.listdir(source_folder_path)
    keys = {}
    origin_keys = {}
    origin_path = ""
    excels_path = []
    for file_name in excel_files:
        path = os.path.join(source_folder_path, file_name)
        if os.path.isfile(path):
            if contain_chinese(file_name) or not (".xlsx" == os.path.splitext(file_name)[-1]):
                print("ignore excel" + file_name)
                continue
            # print(file_name)
            if file_name == "Config.xlsx":
                origin_keys = read_config(path)
                origin_path = path
            else:
                excels_path.append(path)
                dic = read_excel(path)
                if len(dic)>0:
                    keys.update({file_name: dic})

    if len(keys)<=0:
        print("没有需要替换的表格")
        return
    _check_duplication(keys)
    _replace_config_excel(origin_path, keys, origin_keys)
    if len(exception_info)>0:
        for i in range(len(exception_info)):
            print(exception_info[i])
        raise Exception("----------------打表错误-----------------")
    _replace_excels_key(keys)
    replace_translated_excels(keys)
    for i, val in enumerate(excels_path):
        excel_workbook = openpyxl.load_workbook(val)
        sheet_names = excel_workbook.sheetnames
        if len(sheet_names) <= 1:
            continue
        work_sheet = excel_workbook[sheet_names[1]]
        rows = work_sheet.max_row
        for i in range(2, rows + 1):
            work_sheet.delete_rows(2)
        excel_workbook.save(val)
        excel_workbook.close()



def _replace_config_excel(origin_path, keys, origin_keys):
    workbook = openpyxl.load_workbook(origin_path)
    sheet_names = workbook.sheetnames
    work_sheet = workbook[sheet_names[0]]
    rows = work_sheet.max_row
    current_row = rows
    for k, v in keys.items():
        for key, value in v.items():
            if key not in origin_keys:
                error_format = "{0}的key{1}不存在Config.xlsx表格中"
                error_info = error_format.format(k,key)
                exception_info.append(error_info)
                # raise Exception(error_info)
            else:
                row = origin_keys[key][1]
                work_sheet.cell(row,1).value = value[0]
                work_sheet.cell(row, 4).value = k
                work_sheet.cell(row, 5).value = time.strftime("%Y-%m-%d", time.localtime())
                work_sheet.cell(row, 6).value = key
                # old_key = work_sheet.cell(row, 6).value
                # if old_key is not None:
                #     work_sheet.cell(row, 6).value = str(old_key)+","+key
                # else:
                #     work_sheet.cell(row, 6).value = key

    workbook.save(origin_path)
    workbook.close()

def _replace_excels_key(keys):
    replace_dic = {}
    for k, v in keys.items():
        for key, value in v.items():
            replace_dic.update({key: value[0]})

    current_path = os.getcwd()
    source_folder_path = os.path.join(current_path, "../excel")
    excel_files = os.listdir(source_folder_path)
    dirty_excels = []
    for file_name in excel_files:
        file_path = os.path.join(source_folder_path, file_name)
        if contain_chinese(file_name) or not (".xlsx" == os.path.splitext(file_name)[-1]):
            print("ignore excel" + file_name)
        else:
            is_excel_data_dirty = False
            workbook = openpyxl.load_workbook(file_path)
            sheet_names = workbook.sheetnames
            # print(sheet_names[0])
            work_sheet = workbook[sheet_names[0]]
            rows = work_sheet.max_row
            column = work_sheet.max_column
            for col in range(1, column + 1):
                if work_sheet.cell(1, col).value == "string":
                        for row in range(4, rows + 1):
                            key = str(work_sheet.cell(row, col).value)
                            if key in replace_dic:
                                print("替换key="+key+"    filename="+file_name)
                                is_excel_data_dirty = True
                                work_sheet.cell(row, col).value=replace_dic[key]
            if is_excel_data_dirty:
                dirty_excels.append(file_name)
                workbook.save(file_path)
                workbook.close()
    time.sleep(1)
    resave_all_excels(dirty_excels)


def replace_translated_excels(keys):
    print("替换已翻译表格")
    replace_dic = {}
    for k, v in keys.items():
        for key, value in v.items():
            replace_dic.update({key: value[0]})

    current_path = os.getcwd()
    config_path = os.path.join(current_path,"config.txt")

    with open(config_path,encoding='utf-8') as file:
        translated_excel_path =file.readline()
        print(translated_excel_path)
    for path, dir_list, file_list in os.walk(translated_excel_path):
        print(path)
        for file_name in file_list:
            if (file_name != "Config.xlsx"):
                continue
            print(os.path.join(path, file_name))
            file_path = os.path.join(path, file_name)
            is_excel_data_dirty = False
            workbook = openpyxl.load_workbook(file_path)
            sheet_names = workbook.sheetnames
            # print(sheet_names[0])
            work_sheet = workbook[sheet_names[0]]
            rows = work_sheet.max_row
            column = work_sheet.max_column
            for row in range(1, rows + 1):
                key = str(work_sheet.cell(row, 1).value)
                if key in replace_dic:
                    print("替换key=" + key + "    filename=" + file_name)
                    is_excel_data_dirty = True
                    work_sheet.cell(row, 1).value = replace_dic[key]
            if is_excel_data_dirty:
                workbook.save(file_path)
                workbook.close()



def _check_duplication(keys):
    dic = {}
    for k, v in keys.items():
        for key, value in v.items():
            if key in dic:
                format = "{0}和{1}存在重复的key={2}"
                exception_info.append(format.format(k, dic[key], key))
                # raise Exception(format.format(k, dic[key], key))
            dic.update({key: k})


def read_config(file_path):
    # print("read excel " + file_path)
    workbook = openpyxl.load_workbook(file_path)
    sheet_names = workbook.sheetnames
    work_sheet = workbook[sheet_names[0]]
    rows = work_sheet.max_row
    column = work_sheet.max_column
    dic = {}
    for row in range(2, rows + 1):
        show_key = work_sheet.cell(row, 1).value
        content = work_sheet.cell(row, 2).value
        show_key = str(show_key).strip()
        content = str(content)
        if show_key in dic:
            error_info = "key重复 key=" + show_key + " excel=" + file_path
            exception_info.append(error_info)
            # raise Exception(error_info)
        else:
            dic.update({show_key: [content, row]})
    workbook.close()
    return dic


def read_excel(file_path):
    # print("read excel " + file_path)
    workbook = openpyxl.load_workbook(file_path)
    sheet_names = workbook.sheetnames
    sheet_count = len(workbook.sheetnames)
    dic = {}
    if sheet_count <= 1:
        return dic
    work_sheet = workbook[sheet_names[1]]
    rows = work_sheet.max_row
    column = work_sheet.max_column


    for row in range(2, rows + 1):
        replace_key = work_sheet.cell(row, 1).value
        origin_key = work_sheet.cell(row, 2).value
        if replace_key is None or origin_key is None:
            continue
        content = work_sheet.cell(row, 3).value
        replace_key = str(replace_key).strip()
        replace_key = replace_key.replace('\n', '').replace('\r', '')
        origin_key = str(origin_key).strip()
        origin_key = origin_key.replace('\n', '').replace('\r', '')
        content = str(content)
        if origin_key in dic:
            error_info = "key重复 key=" + origin_key + " excel=" + file_path
            exception_info.append(error_info)
            # raise Exception(error_info)
        else:
            dic.update({origin_key: [replace_key, content, row]})
    workbook.close()
    return dic


def resave_all_excels(dirty_excels):
    current_path = os.getcwd()
    source_folder_path = os.path.join(current_path, "../excel")
    excel_files = os.listdir(source_folder_path)
    excel_app = xlwings.App(visible=False)
    print("修改的表格数量="+str(len(dirty_excels)))
    for file_name in excel_files:
        file_path = os.path.join(source_folder_path, file_name)
        if contain_chinese(file_name) or not (".xlsx" == os.path.splitext(file_name)[-1]):
            print("ignore excel" + file_name)
        else:
            if file_name in dirty_excels:
                print("resave excel="+file_path)
                excel_book = excel_app.books.open(file_path)
                excel_book.save()
                excel_book.close()

    excel_app.quit()

if __name__ == "__main__":
    start_replace()


